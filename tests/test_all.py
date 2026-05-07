"""
tests/test_all.py
Полный набор тестов ITIS Platform.
Покрывает: models (itis, bak, ml_predictor) + REST API + Excel-экспорт.
Запуск: pytest tests/ -v
"""
import io, math, pytest, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from models.itis        import Module, calculate, classify, THRESHOLD_LOW, THRESHOLD_HIGH, CLASS_LOW, CLASS_MID, CLASS_HIGH
from models.bak         import BakMatrix, Goal, BakModule
from models.ml_predictor import PredictInput, predict, FEATURE_KEYS, CLASS_HIGH as ML_HIGH, CLASS_MID as ML_MID, CLASS_LOW as ML_LOW
from app import create_app

# ═══════════════════════════════════════════════════════════════════
#  FIXTURES
# ═══════════════════════════════════════════════════════════════════
@pytest.fixture
def client():
    app = create_app(testing=True)
    with app.test_client() as c:
        yield c

@pytest.fixture
def typical_modules():
    return [
        Module("NLP",      E=.95, W=.30, C=.90, M=.85),
        Module("CRM",      E=.88, W=.25, C=.85, M=.90),
        Module("Омниканал",E=.90, W=.20, C=.80, M=.88),
        Module("Аналитика",E=.78, W=.25, C=.75, M=.80),
    ]

@pytest.fixture
def itis_payload():
    return {"modules": [
        {"name":"A","E":.90,"W":.50,"C":.85,"M":.88},
        {"name":"B","E":.75,"W":.50,"C":.70,"M":.80},
    ]}

@pytest.fixture
def bak_matrix():
    bak = BakMatrix(
        goals=[Goal(1,"Рост конверсии","CVR"), Goal(2,"Снижение оттока","Churn")],
        modules=[BakModule(1,"NLP-чат-бот"), BakModule(2,"CRM"), BakModule(3,"Аналитика")]
    )
    bak.set_cell(1,1,True); bak.set_cell(1,2,True)
    bak.set_cell(2,2,True); bak.set_cell(2,3,True)
    return bak

@pytest.fixture
def ml_input_high():
    return PredictInput(digitalization=.95,data_readiness=.90,mgmt_support=.95,
                        goal_clarity=.92,budget=.90,company_size=.95,industry_coeff=.90)

@pytest.fixture
def ml_input_low():
    return PredictInput(digitalization=.10,data_readiness=.10,mgmt_support=.15,
                        goal_clarity=.10,budget=.10,company_size=.50,industry_coeff=.55)

@pytest.fixture
def ml_payload_high():
    return {k: 0.95 for k in FEATURE_KEYS}

@pytest.fixture
def bak_api_payload(bak_matrix):
    return bak_matrix.to_dict()


# ═══════════════════════════════════════════════════════════════════
#  MODEL — ITIS
# ═══════════════════════════════════════════════════════════════════
class TestITISModel:
    def test_itisi_perfect(self):
        assert Module("T",1,1,1,1).itis_i == pytest.approx(1.0, abs=1e-9)

    def test_itisi_zero(self):
        assert Module("T",0,1,0,0).itis_i == pytest.approx(0.0, abs=1e-9)

    def test_itisi_formula(self):
        m = Module("T",.8,.5,.9,.7)
        assert m.itis_i == pytest.approx((.8*.9*.7)**(1/3), rel=1e-9)

    def test_w_not_in_itisi(self):
        assert Module("T",.8,.2,.8,.8).itis_i == pytest.approx(Module("T",.8,.9,.8,.8).itis_i)

    def test_itisi_range(self):
        for e,c,m in [(0,0,0),(1,1,1),(.5,.5,.5),(.3,.7,.9)]:
            assert 0 <= Module("T",e,1,c,m).itis_i <= 1

    def test_classify_low(self):
        assert classify(0.0) == CLASS_LOW
        assert classify(THRESHOLD_LOW - 0.001) == CLASS_LOW

    def test_classify_mid(self):
        assert classify(THRESHOLD_LOW) == CLASS_MID
        assert classify(THRESHOLD_HIGH - 0.001) == CLASS_MID

    def test_classify_high(self):
        assert classify(THRESHOLD_HIGH) == CLASS_HIGH
        assert classify(1.0) == CLASS_HIGH

    def test_single_module(self):
        m = Module("S",.9,1,.85,.88)
        assert calculate([m]).total_itis == pytest.approx(m.itis_i, rel=1e-6)

    def test_equal_weights(self):
        m1, m2 = Module("A",.8,.5,.8,.8), Module("B",.6,.5,.6,.6)
        r = calculate([m1,m2])
        assert r.total_itis == pytest.approx(.5*m1.itis_i + .5*m2.itis_i, rel=1e-6)

    def test_weights_normalized(self):
        r1 = calculate([Module("A",.9,3,.9,.9), Module("B",.9,7,.9,.9)]).total_itis
        r2 = calculate([Module("A",.9,.3,.9,.9),Module("B",.9,.7,.9,.9)]).total_itis
        assert r1 == pytest.approx(r2, rel=1e-6)

    def test_empty_raises(self):
        with pytest.raises(ValueError, match="пустым"): calculate([])

    def test_contributions_sum(self, typical_modules):
        r = calculate(typical_modules)
        assert sum(c["contribution"] for c in r.contributions) == pytest.approx(r.total_itis, rel=1e-4)

    def test_best_worst(self):
        m1, m2 = Module("Best",1,1,1,1), Module("Worst",.1,1,.1,.1)
        r = calculate([m1,m2])
        assert "Best" in r.best_module and "Worst" in r.worst_module

    def test_result_keys(self, typical_modules):
        d = calculate(typical_modules).to_dict()
        for k in ("total_itis","w_sum","efficiency_class","recommendation","best_module","worst_module","modules","contributions"):
            assert k in d

    def test_validate_empty_name(self):
        with pytest.raises(ValueError): Module("  ",.5,.5,.5,.5).validate()

    def test_validate_E_out_of_range(self):
        with pytest.raises(ValueError): Module("T",1.5,.5,.5,.5).validate()

    def test_validate_W_zero(self):
        with pytest.raises(ValueError): Module("T",.5,0,.5,.5).validate()

    def test_from_dict_roundtrip(self):
        m = Module("R",.77,.33,.55,.88)
        r = Module.from_dict(m.to_dict())
        assert r.name == m.name
        assert r.itis_i == pytest.approx(m.itis_i, rel=1e-6)


# ═══════════════════════════════════════════════════════════════════
#  MODEL — BAK MATRIX
# ═══════════════════════════════════════════════════════════════════
class TestBakModel:
    def test_toggle_on(self, bak_matrix):
        prev = bak_matrix.is_covered(1,3)
        bak_matrix.toggle(1,3)
        assert bak_matrix.is_covered(1,3) == (not prev)

    def test_toggle_twice_restores(self, bak_matrix):
        was = bak_matrix.is_covered(1,1)
        bak_matrix.toggle(1,1); bak_matrix.toggle(1,1)
        assert bak_matrix.is_covered(1,1) == was

    def test_set_cell(self, bak_matrix):
        bak_matrix.set_cell(2,1,True)
        assert bak_matrix.is_covered(2,1)
        bak_matrix.set_cell(2,1,False)
        assert not bak_matrix.is_covered(2,1)

    def test_module_coverage_score(self, bak_matrix):
        # Module 2 (CRM) covers both goals
        assert bak_matrix.module_coverage_score(2) == pytest.approx(1.0)
        # Module 1 (NLP) covers only goal 1
        assert bak_matrix.module_coverage_score(1) == pytest.approx(0.5)

    def test_goal_coverage_pct(self, bak_matrix):
        # Goal 1 covered by modules 1,2 out of 3
        assert bak_matrix.goal_coverage_pct(1) == pytest.approx(2/3, rel=1e-4)

    def test_suggested_weights_sum_to_1(self, bak_matrix):
        ws = bak_matrix.suggested_weights()
        assert sum(w["weight"] for w in ws) == pytest.approx(1.0, abs=0.01)

    def test_suggested_weights_highest_for_most_covered(self, bak_matrix):
        ws = {w["module_id"]: w["weight"] for w in bak_matrix.suggested_weights()}
        # Module 2 covers 2 goals, should have highest weight
        assert ws[2] >= ws[1] and ws[2] >= ws[3]

    def test_empty_matrix_weights(self):
        bak = BakMatrix(goals=[Goal(1,"G","")], modules=[BakModule(1,"M")])
        ws = bak.suggested_weights()
        assert len(ws) == 1

    def test_kpi_summary_length(self, bak_matrix):
        assert len(bak_matrix.kpi_summary()) == len(bak_matrix.goals)

    def test_from_dict_roundtrip(self, bak_matrix):
        d = bak_matrix.to_dict()
        restored = BakMatrix.from_dict(d)
        assert restored.is_covered(1,1) == bak_matrix.is_covered(1,1)
        assert restored.is_covered(2,3) == bak_matrix.is_covered(2,3)


# ═══════════════════════════════════════════════════════════════════
#  MODEL — ML PREDICTOR
# ═══════════════════════════════════════════════════════════════════
class TestMLModel:
    def test_high_inputs_give_high_class(self, ml_input_high):
        r = predict(ml_input_high)
        assert r.predicted_class == ML_HIGH

    def test_low_inputs_give_low_class(self, ml_input_low):
        r = predict(ml_input_low)
        assert r.predicted_class == ML_LOW

    def test_probabilities_sum_to_1(self, ml_input_high):
        r = predict(ml_input_high)
        assert sum(r.probabilities.values()) == pytest.approx(1.0, abs=0.01)

    def test_all_probabilities_positive(self, ml_input_high):
        r = predict(ml_input_high)
        assert all(v > 0 for v in r.probabilities.values())

    def test_raw_score_in_range(self, ml_input_high):
        assert 0 <= predict(ml_input_high).raw_score <= 1

    def test_feature_importance_length(self, ml_input_high):
        assert len(predict(ml_input_high).feature_importance) == len(FEATURE_KEYS)

    def test_feature_importance_sums_to_1(self, ml_input_high):
        total = sum(f["importance"] for f in predict(ml_input_high).feature_importance)
        assert total == pytest.approx(1.0, abs=0.01)

    def test_recommendations_not_empty(self, ml_input_high):
        assert len(predict(ml_input_high).recommendations) > 0

    def test_to_dict_keys(self, ml_input_high):
        d = predict(ml_input_high).to_dict()
        for k in ("predicted_class","probabilities","raw_score","feature_importance","feature_contributions","recommendations"):
            assert k in d

    def test_validate_out_of_range(self):
        inp = PredictInput(digitalization=1.5,data_readiness=.5,mgmt_support=.5,
                           goal_clarity=.5,budget=.5,company_size=.5,industry_coeff=.5)
        with pytest.raises(ValueError): inp.validate()

    def test_from_dict_valid(self):
        d = {k: 0.7 for k in FEATURE_KEYS}
        inp = PredictInput.from_dict(d)
        r = predict(inp)
        assert r.predicted_class in (ML_HIGH, ML_MID, ML_LOW)

    def test_from_dict_missing_key(self):
        d = {k: 0.7 for k in FEATURE_KEYS[:-1]}  # missing last key
        with pytest.raises(KeyError): PredictInput.from_dict(d)


# ═══════════════════════════════════════════════════════════════════
#  API — /api/health
# ═══════════════════════════════════════════════════════════════════
class TestHealth:
    def test_global_health(self, client):
        assert client.get("/api/health").status_code == 200

    def test_itis_health(self, client):
        assert client.get("/api/itis/health").status_code == 200

    def test_bak_health(self, client):
        assert client.get("/api/bak/health").status_code == 200

    def test_ml_health(self, client):
        assert client.get("/api/ml/health").status_code == 200

    def test_features_endpoint(self, client):
        r = client.get("/api/ml/features")
        assert r.status_code == 200
        assert "features" in r.get_json()


# ═══════════════════════════════════════════════════════════════════
#  API — /api/itis/calculate
# ═══════════════════════════════════════════════════════════════════
class TestITISAPI:
    def test_valid_200(self, client, itis_payload):
        assert client.post("/api/itis/calculate", json=itis_payload).status_code == 200

    def test_response_total_itis(self, client, itis_payload):
        d = client.post("/api/itis/calculate", json=itis_payload).get_json()
        assert 0 <= d["total_itis"] <= 1

    def test_response_efficiency_class(self, client, itis_payload):
        d = client.post("/api/itis/calculate", json=itis_payload).get_json()
        assert d["efficiency_class"] in (CLASS_LOW, CLASS_MID, CLASS_HIGH)

    def test_high_scores_high_class(self, client):
        p = {"modules":[{"name":"A","E":.99,"W":.5,"C":.99,"M":.99},
                         {"name":"B","E":.98,"W":.5,"C":.98,"M":.98}]}
        assert client.post("/api/itis/calculate",json=p).get_json()["efficiency_class"] == CLASS_HIGH

    def test_low_scores_low_class(self, client):
        p = {"modules":[{"name":"A","E":.1,"W":.5,"C":.1,"M":.1}]}
        assert client.post("/api/itis/calculate",json=p).get_json()["efficiency_class"] == CLASS_LOW

    def test_missing_modules_400(self, client):
        assert client.post("/api/itis/calculate",json={"project":{}}).status_code == 400

    def test_invalid_E_422(self, client):
        assert client.post("/api/itis/calculate",json={"modules":[{"name":"T","E":2,"W":1,"C":.5,"M":.5}]}).status_code == 422

    def test_negative_W_422(self, client):
        assert client.post("/api/itis/calculate",json={"modules":[{"name":"T","E":.5,"W":-1,"C":.5,"M":.5}]}).status_code == 422

    def test_no_body_400(self, client):
        assert client.post("/api/itis/calculate",data="bad",content_type="text/plain").status_code == 400

    def test_validate_valid(self, client):
        d = client.post("/api/itis/validate",json={"name":"T","E":.9,"W":.5,"C":.8,"M":.85}).get_json()
        assert d["valid"] is True

    def test_validate_invalid(self, client):
        d = client.post("/api/itis/validate",json={"name":"T","E":1.5,"W":.5,"C":.5,"M":.5}).get_json()
        assert d["valid"] is False

    def test_contributions_present(self, client, itis_payload):
        d = client.post("/api/itis/calculate",json=itis_payload).get_json()
        assert "contributions" in d
        assert all("share_pct" in c for c in d["contributions"])

    def test_unnormalized_weights_ok(self, client):
        p = {"modules":[{"name":"A","E":.8,"W":3,"C":.8,"M":.8},
                         {"name":"B","E":.8,"W":7,"C":.8,"M":.8}]}
        assert client.post("/api/itis/calculate",json=p).status_code == 200


# ═══════════════════════════════════════════════════════════════════
#  API — /api/bak
# ═══════════════════════════════════════════════════════════════════
class TestBAKAPI:
    def test_weights_200(self, client, bak_api_payload):
        assert client.post("/api/bak/weights",json=bak_api_payload).status_code == 200

    def test_weights_have_sum_1(self, client, bak_api_payload):
        d = client.post("/api/bak/weights",json=bak_api_payload).get_json()
        assert sum(w["weight"] for w in d["weights"]) == pytest.approx(1.0, abs=0.01)

    def test_weights_no_body_400(self, client):
        assert client.post("/api/bak/weights",data="x",content_type="text/plain").status_code == 400

    def test_summary_200(self, client, bak_api_payload):
        assert client.post("/api/bak/summary",json=bak_api_payload).status_code == 200

    def test_summary_has_matrix(self, client, bak_api_payload):
        d = client.post("/api/bak/summary",json=bak_api_payload).get_json()
        assert "matrix" in d and "kpi_summary" in d

    def test_toggle_200(self, client, bak_api_payload):
        payload = {"goal_id":1,"module_id":3,"matrix":bak_api_payload}
        r = client.post("/api/bak/toggle",json=payload)
        assert r.status_code == 200
        d = r.get_json()
        assert "covered" in d and "weights" in d


# ═══════════════════════════════════════════════════════════════════
#  API — /api/ml/predict
# ═══════════════════════════════════════════════════════════════════
class TestMLAPI:
    def test_high_payload_200(self, client, ml_payload_high):
        assert client.post("/api/ml/predict",json=ml_payload_high).status_code == 200

    def test_high_payload_class(self, client, ml_payload_high):
        d = client.post("/api/ml/predict",json=ml_payload_high).get_json()
        assert d["predicted_class"] == ML_HIGH

    def test_low_payload_class(self, client):
        p = {k: 0.05 for k in FEATURE_KEYS}
        d = client.post("/api/ml/predict",json=p).get_json()
        assert d["predicted_class"] == ML_LOW

    def test_probabilities_present(self, client, ml_payload_high):
        d = client.post("/api/ml/predict",json=ml_payload_high).get_json()
        assert "probabilities" in d and len(d["probabilities"]) == 3

    def test_missing_key_400(self, client):
        p = {k: 0.7 for k in FEATURE_KEYS[:-1]}
        assert client.post("/api/ml/predict",json=p).status_code == 400

    def test_out_of_range_422(self, client):
        p = {k: 0.7 for k in FEATURE_KEYS}; p[FEATURE_KEYS[0]] = 2.0
        assert client.post("/api/ml/predict",json=p).status_code == 422

    def test_no_body_400(self, client):
        assert client.post("/api/ml/predict",data="bad",content_type="text/plain").status_code == 400

    def test_recommendations_in_response(self, client, ml_payload_high):
        d = client.post("/api/ml/predict",json=ml_payload_high).get_json()
        assert len(d["recommendations"]) > 0

    def test_feature_contributions_present(self, client, ml_payload_high):
        d = client.post("/api/ml/predict",json=ml_payload_high).get_json()
        assert "feature_contributions" in d


# ═══════════════════════════════════════════════════════════════════
#  EXPORT — Excel files
# ═══════════════════════════════════════════════════════════════════
class TestExport:
    def _is_xlsx(self, data: bytes) -> bool:
        return data[:2] == b"PK"

    def test_itis_export_200(self, client, itis_payload):
        assert client.post("/export/itis",json=itis_payload).status_code == 200

    def test_itis_export_xlsx(self, client, itis_payload):
        r = client.post("/export/itis",json=itis_payload)
        assert self._is_xlsx(r.data)

    def test_itis_export_readable(self, client, itis_payload):
        import openpyxl
        r = client.post("/export/itis",json=itis_payload)
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert "Результаты ITIS" in wb.sheetnames
        assert "Методика ITIS"   in wb.sheetnames

    def test_itis_export_no_modules_400(self, client):
        assert client.post("/export/itis",json={"project":{}}).status_code == 400

    def test_bak_export_200(self, client, bak_api_payload):
        assert client.post("/export/bak",json=bak_api_payload).status_code == 200

    def test_bak_export_xlsx(self, client, bak_api_payload):
        r = client.post("/export/bak",json=bak_api_payload)
        assert self._is_xlsx(r.data)

    def test_bak_export_two_sheets(self, client, bak_api_payload):
        import openpyxl
        r = client.post("/export/bak",json=bak_api_payload)
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert "БАК-матрица" in wb.sheetnames
        assert "Рекомендованные веса" in wb.sheetnames

    def test_ml_export_200(self, client, ml_payload_high):
        assert client.post("/export/ml",json=ml_payload_high).status_code == 200

    def test_ml_export_xlsx(self, client, ml_payload_high):
        r = client.post("/export/ml",json=ml_payload_high)
        assert self._is_xlsx(r.data)

    def test_ml_export_two_sheets(self, client, ml_payload_high):
        import openpyxl
        r = client.post("/export/ml",json=ml_payload_high)
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert "ML-прогноз" in wb.sheetnames
        assert "Рекомендации" in wb.sheetnames

    def test_ml_export_missing_key_400(self, client):
        p = {k: 0.7 for k in FEATURE_KEYS[:-1]}
        assert client.post("/export/ml",json=p).status_code == 400

    def test_index_page_200(self, client):
        assert client.get("/").status_code == 200
