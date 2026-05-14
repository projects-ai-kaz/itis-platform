"""routes/export_companies.py — Экспорт данных компаний в Excel."""

from __future__ import annotations

import io
from flask import Blueprint, jsonify, request, send_file

export_companies_bp = Blueprint("export_companies", __name__, url_prefix="/export/companies")


@export_companies_bp.post("/")
def export_companies():
    """
    POST /export/companies/
    Body (опционально): {"ids": ["uuid1", ...]}  — если пусто, экспортируются все компании.
    Возвращает Excel-файл с двумя листами:
      1. Все компании (сводная таблица)
      2. Модули (детализация)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl не установлен. pip install openpyxl"}), 500

    from models.company_store import store

    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])
    if ids:
        companies = [store.get_company(cid) for cid in ids if store.get_company(cid)]
    else:
        companies = store.list_companies()

    if not companies:
        return jsonify({"error": "Нет данных для экспорта"}), 404

    wb = openpyxl.Workbook()

    # ── Лист 1: Сводная таблица компаний ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Компании"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    alt_fill = PatternFill("solid", fgColor="EEF4FB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    headers1 = ["#", "Компания", "Отрасль", "Страна", "Кол-во модулей", "ITIS Score", "Класс эффективности", "Создана"]
    ws1.append(headers1)
    for col_idx, _ in enumerate(headers1, 1):
        cell = ws1.cell(1, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Сортируем по ITIS desc
    companies_sorted = sorted(companies, key=lambda c: c.get("itis_score", 0), reverse=True)

    for row_idx, cp in enumerate(companies_sorted, 2):
        row_data = [
            row_idx - 1,
            cp.get("name", ""),
            cp.get("industry", ""),
            cp.get("country", ""),
            cp.get("modules_count", len(cp.get("modules", []))),
            cp.get("itis_score", 0),
            cp.get("itis_class", ""),
            cp.get("created_at", "")[:10],
        ]
        ws1.append(row_data)
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, _ in enumerate(row_data, 1):
            cell = ws1.cell(row_idx, col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            if fill:
                cell.fill = fill
            # Цвет ITIS
            if col_idx == 6:
                score = cp.get("itis_score", 0)
                if score >= 0.70:
                    cell.font = Font(color="1A7A4A", bold=True)
                elif score >= 0.45:
                    cell.font = Font(color="B07D0A", bold=True)
                else:
                    cell.font = Font(color="C0392B", bold=True)

    # Ширина колонок
    col_widths1 = [5, 30, 20, 15, 18, 14, 28, 14]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 22

    # ── Лист 2: Детализация по модулям ───────────────────────────────────
    ws2 = wb.create_sheet("Модули")
    headers2 = ["Компания", "Отрасль", "Модуль", "E (Эффект)", "W (Вес)", "C (Покрытие)", "M (Зрелость)", "ITISᵢ"]
    ws2.append(headers2)
    for col_idx, _ in enumerate(headers2, 1):
        cell = ws2.cell(1, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = 2
    for cp in companies_sorted:
        for mod in cp.get("modules", []):
            e, w, c, m = float(mod.get("E", 0)), float(mod.get("W", 0)), float(mod.get("C", 0)), float(mod.get("M", 0))
            product = max(0.0, e * c * m)
            itis_i = round(product ** (1 / 3), 4)
            ws2.append([
                cp.get("name", ""),
                cp.get("industry", ""),
                mod.get("name", ""),
                e, w, c, m, itis_i
            ])
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, 9):
                cell = ws2.cell(row_idx, col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx not in (1, 2, 3) else "left")
                if fill:
                    cell.fill = fill
            row_idx += 1

    col_widths2 = [28, 18, 24, 14, 12, 14, 14, 12]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 22

    # Freeze headers
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="companies_itis_report.xlsx"
    )
