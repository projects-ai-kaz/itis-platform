"""routes/export.py — Excel-экспорт для всех трёх вкладок."""
import io
from datetime import date
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.itis import Module, calculate, classify
from models.bak import BakMatrix
from models.ml_predictor import PredictInput, predict

export_bp = Blueprint("export", __name__, url_prefix="/export")

# ── Palette ────────────────────────────────────────────────────────────────
BG   = "080b10"; BG2 = "0e1219"; BG3 = "141a24"
ACC  = "00E5B4"; ACC2 = "3B82F6"; ACC3 = "F59E0B"
RED  = "EF4444"; TXT  = "DDE2ED"; MUTE = "7E8BA0"

def _s(color="1C2332", style="thin"):
    return Side(border_style=style, color=color)

def _brd():
    s = _s(); return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, r, c, v="", bold=False, color=TXT, bg=None, size=10,
          align="left", wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = _brd()
    return cell

def _cls_color(cls: str) -> str:
    if "Низ" in cls or "Low" in cls: return RED
    if "Сред" in cls or "Mid" in cls: return ACC3
    return ACC

def _to_bytes(wb) -> bytes:
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ══════════════════════════════════════════════
#  ITIS export
# ══════════════════════════════════════════════
def _build_itis_xlsx(project: dict, modules: list, result) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "Результаты ITIS"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [3,28,8,8,8,8,10,12,26]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:I2")
    c = ws["B2"]; c.value = "ОТЧЁТ — AI IMPACT SCORE (ITIS)"
    c.font = Font(name="Calibri", bold=True, color=ACC, size=15)
    c.fill = PatternFill("solid", fgColor=BG); c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 34

    info = [("Компания", project.get("company","—")), ("AI-решение", project.get("solution","—")),
            ("Аналитик", project.get("analyst","—")), ("Дата", project.get("eval_date", str(date.today())))]
    for i, (lbl, val) in enumerate(info, 4):
        _cell(ws, i, 2, lbl, color=MUTE, bg=BG2, size=9)
        ws.merge_cells(f"C{i}:I{i}"); _cell(ws, i, 3, val, bold=True, bg=BG2)
        ws.row_dimensions[i].height = 18

    ws.merge_cells("B9:E9"); _cell(ws, 9, 2, "ИТОГОВЫЙ ITIS", color=MUTE, bg=BG3, bold=True, size=9)
    ws.merge_cells("F9:G9"); _cell(ws, 9, 6, round(result.total_itis,4), bold=True, bg=BG3,
                                   color=_cls_color(result.efficiency_class), size=13, align="center")
    ws.merge_cells("H9:I9"); _cell(ws, 9, 8, result.efficiency_class, bold=True,
                                   color=_cls_color(result.efficiency_class), bg=BG3)
    ws.row_dimensions[9].height = 26

    hdrs = ["№","Модуль","E","W","C","M","ITISᵢ","Класс","Рекомендация"]
    for j, h in enumerate(hdrs, 2):
        _cell(ws, 11, j, h, bold=True, color=ACC, bg=BG, align="center", size=9)
    ws.row_dimensions[11].height = 20

    for idx, m in enumerate(modules, 1):
        r = 11+idx; rb = BG3 if idx%2 else BG2
        iv = round(m.itis_i, 4); cls = classify(iv)
        vals = [idx, m.name, m.E, m.W, m.C, m.M, iv, cls,
                "Доработать" if "Низ" in cls else ("Оптимизировать" if "Сред" in cls else "Масштабировать")]
        for j, v in enumerate(vals, 2):
            al = "center" if j in (3,4,5,6,7) else "left"
            _cell(ws, r, j, v, bold=j==8, color=_cls_color(cls) if j==8 else TXT, bg=rb, size=9, align=al)
        ws.row_dimensions[r].height = 18

    # Sheet 2 — Methodology
    ws2 = wb.create_sheet("Методика ITIS"); ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["B"].width = 26; ws2.column_dimensions["C"].width = 50; ws2.column_dimensions["D"].width = 14
    ws2.merge_cells("B2:D2"); c2 = ws2["B2"]; c2.value = "Методика AI Impact Score (ITIS)"
    c2.font = Font(name="Calibri", bold=True, color=ACC, size=13)
    c2.fill = PatternFill("solid", fgColor=BG); c2.alignment = Alignment(horizontal="left", vertical="center")
    params = [("E","Бизнес-эффект модуля","[0,1]"),("W","Стратегический вес (ΣW=1)","(>0)"),
              ("C","Покрытие бизнес-целей","[0,1]"),("M","Зрелость реализации (TRL)","[0,1]")]
    for j, h in enumerate(["Параметр","Описание","Диапазон"], 2):
        _cell(ws2, 4, j, h, bold=True, color=ACC, bg=BG, size=9)
    for i,(p,d,rng) in enumerate(params,5):
        bg = BG3 if i%2 else BG2
        _cell(ws2,i,2,p,bold=True,bg=bg); _cell(ws2,i,3,d,color=MUTE,bg=bg,wrap=True); _cell(ws2,i,4,rng,color=ACC,bg=bg,align="center")
        ws2.row_dimensions[i].height = 18
    for i,(lbl,f) in enumerate([("Частный:","ITISᵢ = ∛(E × C × M)"),("Итоговый:","ITIS = Σ(Wᵢ/ΣW × ITISᵢ)")],10):
        _cell(ws2,i,2,lbl,bold=True,color=MUTE,bg=BG3); ws2.merge_cells(f"C{i}:D{i}"); _cell(ws2,i,3,f,bold=True,color=ACC,bg=BG3); ws2.row_dimensions[i].height = 20
    for j,h in enumerate(["ITIS","Класс","Рекомендация"],2): _cell(ws2,13,j,h,bold=True,color=ACC,bg=BG,size=9)
    for i,(rng,cls,rec,col) in enumerate([("0.00–0.44","Низкая эффективность","Пересмотр стратегии",RED),
                                           ("0.45–0.69","Средняя эффективность","Доработка модулей",ACC3),
                                           ("0.70–1.00","Высокая эффективность","Масштабирование",ACC)],14):
        bg = BG3 if i%2 else BG2
        _cell(ws2,i,2,rng,bold=True,color=col,bg=bg,align="center"); _cell(ws2,i,3,cls,bold=True,color=col,bg=bg); _cell(ws2,i,4,rec,color=MUTE,bg=bg,wrap=True)
        ws2.row_dimensions[i].height = 18
    return _to_bytes(wb)


@export_bp.route("/itis", methods=["POST"])
def export_itis():
    data = request.get_json(silent=True)
    if not data: return jsonify({"error": "JSON обязателен."}), 400
    raw = data.get("modules")
    if not raw: return jsonify({"error": "Поле 'modules' обязательно."}), 400
    try:
        modules = [Module.from_dict(m) for m in raw]
        result  = calculate(modules)
        xlsx    = _build_itis_xlsx(data.get("project", {}), modules, result)
    except (KeyError, TypeError) as e: return jsonify({"error": str(e)}), 400
    except ValueError as e:           return jsonify({"error": str(e)}), 422
    co = data.get("project", {}).get("company", "Report").replace(" ","_")
    return send_file(io.BytesIO(xlsx),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"ITIS_{co}.xlsx")


# ══════════════════════════════════════════════
#  BAK export
# ══════════════════════════════════════════════
def _build_bak_xlsx(bak: BakMatrix) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "БАК-матрица"; ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:Z2"); c=ws["B2"]; c.value="БАК — Матрица покрытия «Бизнес-цель × AI-модуль»"
    c.font=Font(name="Calibri",bold=True,color=ACC2,size=13); c.fill=PatternFill("solid",fgColor=BG)
    c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[2].height=30

    # Headers
    _cell(ws,4,2,"Бизнес-цель",bold=True,color=ACC2,bg=BG,size=9,align="center")
    for j,m in enumerate(bak.modules,3):
        _cell(ws,4,j,m.name[:18],bold=True,color=ACC2,bg=BG,size=9,align="center")
        ws.column_dimensions[chr(64+j)].width=16
    _cell(ws,4,len(bak.modules)+3,"% покрытия",bold=True,color=ACC2,bg=BG,size=9,align="center")
    ws.column_dimensions["B"].width=30; ws.row_dimensions[4].height=20

    for ri,g in enumerate(bak.goals,5):
        rb=BG3 if ri%2 else BG2
        _cell(ws,ri,2,g.name,bg=rb,size=9)
        for j,m in enumerate(bak.modules,3):
            v="✓" if bak.is_covered(g.id,m.id) else ""
            col=ACC if v else MUTE
            _cell(ws,ri,j,v,bold=bool(v),color=col,bg=rb,size=11,align="center")
        pct=f"{bak.goal_coverage_pct(g.id)*100:.0f}%"
        _cell(ws,ri,len(bak.modules)+3,pct,color=ACC3,bg=rb,align="center",bold=True,size=9)
        ws.row_dimensions[ri].height=18

    # Weights sheet
    ws2=wb.create_sheet("Рекомендованные веса"); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["B"].width=26; ws2.column_dimensions["C"].width=14; ws2.column_dimensions["D"].width=18
    ws2.merge_cells("B2:D2"); c2=ws2["B2"]; c2.value="Рекомендованные веса W для ITIS-калькулятора"
    c2.font=Font(name="Calibri",bold=True,color=ACC2,size=12); c2.fill=PatternFill("solid",fgColor=BG)
    c2.alignment=Alignment(horizontal="left",vertical="center"); ws2.row_dimensions[2].height=28
    for j,h in enumerate(["Модуль","Вес W","Покрытие целей (%)"],2): _cell(ws2,4,j,h,bold=True,color=ACC2,bg=BG,size=9)
    for i,w in enumerate(bak.suggested_weights(),5):
        bg=BG3 if i%2 else BG2
        _cell(ws2,i,2,w["module_name"],bg=bg); _cell(ws2,i,3,w["weight"],bold=True,color=ACC,bg=bg,align="center")
        _cell(ws2,i,4,f"{w['coverage_score']*100:.1f}%",color=ACC3,bg=bg,align="center")
        ws2.row_dimensions[i].height=18
    return _to_bytes(wb)


@export_bp.route("/bak", methods=["POST"])
def export_bak():
    data = request.get_json(silent=True)
    if not data: return jsonify({"error": "JSON обязателен."}), 400
    try:
        bak  = BakMatrix.from_dict(data)
        xlsx = _build_bak_xlsx(bak)
    except Exception as e: return jsonify({"error": str(e)}), 400
    return send_file(io.BytesIO(xlsx),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="BAK_Matrix.xlsx")


# ══════════════════════════════════════════════
#  ML export
# ══════════════════════════════════════════════
def _build_ml_xlsx(inp: PredictInput, result) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "ML-прогноз"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 32; ws.column_dimensions["C"].width = 16; ws.column_dimensions["D"].width = 28

    ws.merge_cells("B2:D2"); c=ws["B2"]; c.value="ML-прогноз класса эффекта AI-внедрения"
    c.font=Font(name="Calibri",bold=True,color=ACC3,size=13); c.fill=PatternFill("solid",fgColor=BG)
    c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[2].height=30

    _cell(ws,4,2,"Прогнозируемый класс",bold=True,color=MUTE,bg=BG3,size=9)
    _cell(ws,4,3,result.predicted_class,bold=True,color=_cls_color(result.predicted_class),bg=BG3,size=11)
    ws.merge_cells("C4:D4"); ws.row_dimensions[4].height=24

    cls_map = {"Высокий эффект": ACC, "Средний эффект": ACC3, "Низкий эффект": RED}
    for i,(cls,prob) in enumerate(result.probabilities.items(),6):
        bg=BG3 if i%2 else BG2
        _cell(ws,i,2,cls,bg=bg,size=9); _cell(ws,i,3,f"{prob*100:.1f}%",bold=True,color=cls_map.get(cls,TXT),bg=bg,align="center")
        ws.row_dimensions[i].height=18

    _cell(ws,9,2,"Важность признаков",bold=True,color=MUTE,bg=BG,size=9)
    _cell(ws,9,3,"Важность (%)",bold=True,color=ACC3,bg=BG,size=9,align="center")
    _cell(ws,9,4,"Вклад в прогноз",bold=True,color=ACC3,bg=BG,size=9,align="center")
    for i,f in enumerate(result.feature_importance,10):
        bg=BG3 if i%2 else BG2
        _cell(ws,i,2,f["feature"],bg=bg,size=9)
        _cell(ws,i,3,f"{f['importance']*100:.0f}%",color=ACC3,bg=bg,align="center",bold=True)
        contrib_row=next((x for x in result.feature_contributions if x["feature"]==f["feature"]),{})
        _cell(ws,i,4,contrib_row.get("contribution",""),color=ACC,bg=bg,align="center")
        ws.row_dimensions[i].height=18

    ws2=wb.create_sheet("Рекомендации"); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["B"].width=60
    ws2.merge_cells("B2:B2"); c2=ws2["B2"]; c2.value=f"Рекомендации для класса: {result.predicted_class}"
    c2.font=Font(name="Calibri",bold=True,color=_cls_color(result.predicted_class),size=12)
    c2.fill=PatternFill("solid",fgColor=BG); c2.alignment=Alignment(horizontal="left",vertical="center"); ws2.row_dimensions[2].height=28
    for i,rec in enumerate(result.recommendations,4):
        _cell(ws2,i,2,f"{i-3}. {rec}",color=TXT,bg=BG3 if i%2 else BG2,wrap=True,size=10)
        ws2.row_dimensions[i].height=30
    return _to_bytes(wb)


@export_bp.route("/ml", methods=["POST"])
def export_ml():
    data = request.get_json(silent=True)
    if not data: return jsonify({"error": "JSON обязателен."}), 400
    try:
        inp    = PredictInput.from_dict(data)
        result = predict(inp)
        xlsx   = _build_ml_xlsx(inp, result)
    except (KeyError, TypeError) as e: return jsonify({"error": str(e)}), 400
    except ValueError as e:           return jsonify({"error": str(e)}), 422
    return send_file(io.BytesIO(xlsx),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="ML_Predict.xlsx")
